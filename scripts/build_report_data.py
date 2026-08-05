# -*- coding: utf-8 -*-
"""从 budu OS文档 的 Excel 报表生成 Dashboard 数据模块。

用法: python scripts/build_report_data.py
输入: Desktop/budu OS文档/ 下的
  - 三店4月份.xlsx ~ 三店7月份.xlsx（综合营业统计 sheet）
  - 三店4月菜品明细.xlsx ~ 三店7月菜品明细.xlsx（菜品销售统计 sheet）
  - budu薪资表(总）.xlsx（2026.27周 ~ 2026.31周）
输出: src/data/reportData.js
"""
import io
import os
import re
import warnings

import openpyxl

warnings.filterwarnings('ignore')

DOC_DIR = r'C:\Users\Administrator\Desktop\budu OS文档'
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'src', 'data', 'reportData.js')

MONTH_FILES = [
    ('2026-04', '三店4月份.xlsx'),
    ('2026-05', '三店5月份.xlsx'),
    ('2026-06', '三店6月份.xlsx'),
    ('2026-07', '三店7月份.xlsx'),
]

PRODUCT_FILES = [
    ('2026-04', '三店4月菜品明细.xlsx'),
    ('2026-05', '三店5月菜品明细.xlsx'),
    ('2026-06', '三店6月菜品明细.xlsx'),
    ('2026-07', '三店7月菜品明细.xlsx'),
]

STORE_MAP = [
    ('budu（三里屯通盈店）', 'tongying', '北京通盈中心店', '朝阳区 · 三里屯'),
    ('budu（官舍店）', 'guanshe', '北京官舍店', '朝阳区 · 亮马桥'),
    ('budu（西单更新场店）', 'xidan', '北京西单店', '西城区 · 西单'),
]
SOURCE_TO_KEY = {s: k for s, k, _, _ in STORE_MAP}
STORE_EXTRA = [('chaowai', '北京朝外店', '朝阳区 · 朝外大街')]
NAME_MAP = {k: n for _, k, n, _ in STORE_MAP}
NAME_MAP.update({k: n for k, n, _ in STORE_EXTRA})

SALARY_STORE_MAP = {
    '通盈': 'tongying',
    '官舍': 'guanshe',
    '西单': 'xidan',
    '朝外': 'chaowai',
    '大族': 'chaowai',
}

FULL_TIME = {'李飞燕', '叶芷辰', '隋晓'}


def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '')
    if not s or s in ('--', '-'):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date(v):
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    m = re.match(r'(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})', s)
    if m:
        return '%s-%02d-%02d' % (m.group(1), int(m.group(2)), int(m.group(3)))
    return ''


def load_monthly(month_key, fname):
    wb = openpyxl.load_workbook(os.path.join(DOC_DIR, fname), data_only=True)
    ws = wb['综合营业统计']
    daily = {}
    for r in range(6, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if b is None or str(b).strip() == '--' or '合计' in str(b):
            continue
        key = SOURCE_TO_KEY.get(str(b).strip())
        if key is None:
            continue
        d = parse_date(ws.cell(row=r, column=3).value)
        if not d:
            continue
        row = {
            'd': d[5:],
            'rev': round(num(ws.cell(row=r, column=6).value), 2),
            'inc': round(num(ws.cell(row=r, column=8).value), 2),
            'dis': round(num(ws.cell(row=r, column=7).value), 2),
            'ord': int(num(ws.cell(row=r, column=9).value)),
            'dish': int(num(ws.cell(row=r, column=33).value)),
            'dishRev': round(num(ws.cell(row=r, column=30).value), 2),
            'dishInc': round(num(ws.cell(row=r, column=31).value), 2),
            'boxFee': round(num(ws.cell(row=r, column=35).value), 2),
            'inStore': round(num(ws.cell(row=r, column=13).value), 2),
            'mt': round(num(ws.cell(row=r, column=19).value), 2),
            'tb': round(num(ws.cell(row=r, column=25).value), 2),
            'cash': round(num(ws.cell(row=r, column=38).value), 2),
            'wechat': round(num(ws.cell(row=r, column=39).value), 2),
            'alipay': round(num(ws.cell(row=r, column=40).value), 2),
            'union': round(num(ws.cell(row=r, column=41).value), 2),
            'mtPay': round(num(ws.cell(row=r, column=43).value), 2),
            'tbPay': round(num(ws.cell(row=r, column=44).value), 2),
        }
        daily.setdefault(key, []).append(row)
    return daily


def load_products(month_key, fname):
    wb = openpyxl.load_workbook(os.path.join(DOC_DIR, fname), data_only=True)
    ws = wb['菜品销售统计']
    prods = {}
    for r in range(4, ws.max_row + 1):
        store = ws.cell(row=r, column=2).value
        name = ws.cell(row=r, column=3).value
        if not store or not name:
            continue
        key = SOURCE_TO_KEY.get(str(store).strip())
        if key is None:
            continue
        item = {
            'name': str(name).strip(),
            'sales': round(num(ws.cell(row=r, column=4).value), 1),
            'amount': round(num(ws.cell(row=r, column=6).value), 2),
            'income': round(num(ws.cell(row=r, column=8).value), 2),
            'discount': round(num(ws.cell(row=r, column=10).value), 2),
        }
        prods.setdefault(key, []).append(item)
    for k in prods:
        merged = {}
        for p in prods[k]:
            m = merged.setdefault(p['name'], {'name': p['name'], 'sales': 0.0, 'amount': 0.0, 'income': 0.0, 'discount': 0.0})
            for f in ('sales', 'amount', 'income', 'discount'):
                m[f] += p[f]
        prods[k] = sorted(merged.values(), key=lambda x: -x['amount'])
    return prods


WEEK_RE = re.compile(r'(\d+)周')

# ISO 2026 周 -> 所属月份天数（周一开始）
WEEK_MONTH_SPLIT = {
    27: {'2026-06': 2, '2026-07': 5},
    28: {'2026-07': 7},
    29: {'2026-07': 7},
    30: {'2026-07': 7},
    31: {'2026-07': 5, '2026-08': 2},
}


def _new_week():
    return {
        'salary': 0.0, 'baseHours': 0.0, 'otHours': 0.0, 'otPay': 0.0,
        'perf': 0.0, 'big': 0.0, 'holiday': 0.0,
        'workedRevenue': 0.0, 'workedDays': 0, 'achieve': 0, 'duty': 0, 'review': 0,
    }


def split_salary_by_month(employees):
    """将 5 周薪资数据按天数比例拆分到 2026-06 / 2026-07 / 2026-08。"""
    monthly = {}
    for e in employees:
        for wnum, wk in e.get('weeks', {}).items():
            for mk, days in WEEK_MONTH_SPLIT.get(wnum, {}).items():
                ratio = days / 7.0
                rec = monthly.setdefault(mk, {}).setdefault(e['name'], {
                    'name': e['name'],
                    'salary': 0.0, 'baseHours': 0.0, 'otHours': 0.0, 'otPay': 0.0,
                    'perf': 0.0, 'big': 0.0, 'holiday': 0.0,
                    'workedRevenue': 0.0, 'workedDays': 0.0,
                    'achieve': 0.0, 'duty': 0.0, 'review': 0.0,
                })
                for f in ('salary', 'baseHours', 'otHours', 'otPay', 'perf', 'big', 'holiday',
                          'workedRevenue', 'achieve', 'duty', 'review', 'workedDays'):
                    rec[f] += wk[f] * ratio
    out = {}
    for mk in sorted(monthly.keys()):
        out[mk] = []
        for e in employees:
            rec = monthly[mk].get(e['name'])
            if rec is None:
                continue
            sk = pick_store(e)
            out[mk].append({
                'name': e['name'],
                'type': 'fulltime' if e['name'] in FULL_TIME else 'parttime',
                'storeKey': sk,
                'storeName': NAME_MAP.get(sk, '多店支援'),
                'salary': round(rec['salary'], 2),
                'baseHours': round(rec['baseHours'], 1),
                'otHours': round(rec['otHours'], 1),
                'otPay': round(rec['otPay'], 2),
                'perf': round(rec['perf'], 2),
                'big': round(rec['big'], 2),
                'workedRevenue': round(rec['workedRevenue'], 2),
                'workedDays': int(round(rec['workedDays'])),
                'achieve': int(round(rec['achieve'])),
                'duty': int(round(rec['duty'])),
                'review': int(round(rec['review'])),
            })
        out[mk].sort(key=lambda x: -x['salary'])
    return out


def load_salary():
    wb = openpyxl.load_workbook(os.path.join(DOC_DIR, 'budu薪资表(总）.xlsx'), data_only=True)
    emp = {}
    for ws in wb.worksheets:
        m = WEEK_RE.search(ws.title)
        wnum = int(m.group(1)) if m else 0
        cur = None
        for r in range(2, ws.max_row + 1):
            seq = ws.cell(row=r, column=2).value
            name = ws.cell(row=r, column=3).value
            if seq is not None and name is not None:
                cur = emp.setdefault(str(name).strip(), {
                    'name': str(name).strip(),
                    'stores': {},
                    'weeks': {},
                    'salary': 0.0, 'baseHours': 0.0, 'otHours': 0.0, 'otPay': 0.0,
                    'perf': 0.0, 'big': 0.0, 'holiday': 0.0,
                    'workedRevenue': 0.0, 'workedDays': 0, 'achieve': 0, 'duty': 0, 'review': 0,
                })
            if cur is None:
                continue
            wk = cur['weeks'].setdefault(wnum, _new_week())
            store_raw = ws.cell(row=r, column=7).value
            if store_raw:
                store_raw = str(store_raw).strip()
                if store_raw not in ('休息',):
                    cur['stores'][store_raw] = cur['stores'].get(store_raw, 0) + 1
            shift = ws.cell(row=r, column=8).value
            shift = str(shift).strip() if shift else ''
            if ws.cell(row=r, column=4).value == '合计':
                vals = {
                    'salary': num(ws.cell(row=r, column=21).value),
                    'baseHours': num(ws.cell(row=r, column=13).value),
                    'otHours': num(ws.cell(row=r, column=16).value),
                    'otPay': num(ws.cell(row=r, column=17).value),
                    'perf': num(ws.cell(row=r, column=18).value),
                    'big': num(ws.cell(row=r, column=19).value),
                    'holiday': num(ws.cell(row=r, column=20).value),
                }
                for k, v in vals.items():
                    cur[k] += v
                    wk[k] += v
            else:
                rev = num(ws.cell(row=r, column=9).value)
                if shift in ('早班', '晚班', '中班'):
                    cur['workedDays'] += 1
                    cur['workedRevenue'] += rev
                    wk['workedDays'] += 1
                    wk['workedRevenue'] += rev
                if ws.cell(row=r, column=10).value == '✅':
                    cur['achieve'] += 1
                    wk['achieve'] += 1
                if num(ws.cell(row=r, column=11).value) > 0:
                    cur['duty'] += 1
                    wk['duty'] += 1
                if ws.cell(row=r, column=12).value == '✅':
                    cur['review'] += 1
                    wk['review'] += 1
    return list(emp.values())

def pick_store(emp):
    if not emp['stores']:
        return 'multi'
    top = max(emp['stores'].items(), key=lambda kv: (kv[1], list(emp['stores']).index(kv[0])))
    return SALARY_STORE_MAP.get(top[0], 'multi')


def main():
    import json

    all_daily = {}
    monthly = {}
    for month_key, fname in MONTH_FILES:
        daily = load_monthly(month_key, fname)
        all_daily[month_key] = daily
        for key in SOURCE_TO_KEY.values():
            rows = daily.get(key, [])
            agg = {
                'days': len(rows),
                'rev': round(sum(x['rev'] for x in rows), 2),
                'inc': round(sum(x['inc'] for x in rows), 2),
                'dis': round(sum(x['dis'] for x in rows), 2),
                'ord': sum(x['ord'] for x in rows),
                'dish': sum(x['dish'] for x in rows),
                'dishRev': round(sum(x['dishRev'] for x in rows), 2),
                'dishInc': round(sum(x['dishInc'] for x in rows), 2),
                'boxFee': round(sum(x['boxFee'] for x in rows), 2),
                'inStore': round(sum(x['inStore'] for x in rows), 2),
                'mt': round(sum(x['mt'] for x in rows), 2),
                'tb': round(sum(x['tb'] for x in rows), 2),
            }
            monthly.setdefault(key, {})[month_key] = agg

    products = {}
    for month_key, fname in PRODUCT_FILES:
        products[month_key] = load_products(month_key, fname)

    employees = load_salary()
    emp_out = []
    for e in employees:
        sk = pick_store(e)
        emp_out.append({
            'name': e['name'],
            'type': 'fulltime' if e['name'] in FULL_TIME else 'parttime',
            'storeKey': sk,
            'storeName': NAME_MAP.get(sk, '多店支援'),
            'salary': round(e['salary'], 2),
            'baseHours': round(e['baseHours'], 1),
            'otHours': round(e['otHours'], 1),
            'otPay': round(e['otPay'], 2),
            'perf': round(e['perf'], 2),
            'big': round(e['big'], 2),
            'workedRevenue': round(e['workedRevenue'], 2),
            'workedDays': e['workedDays'],
            'achieve': e['achieve'],
            'duty': e['duty'],
            'review': e['review'],
        })
    
    emp_monthly = split_salary_by_month(employees)
    emp_out.sort(key=lambda x: -x['salary'])

    lines = []
    lines.append('// 本文件由 scripts/build_report_data.py 自动生成，请勿手动修改。')
    lines.append('// 数据来源：Desktop/budu OS文档/（三店4-7月报表 + 菜品明细 + 薪资表(总)）')
    lines.append('')
    lines.append('export const DATA_SOURCE = %s' % json.dumps('budu OS文档 · 三店4-7月报表 / 菜品明细 / 薪资表（2026.27-31周）', ensure_ascii=False))
    lines.append('')
    lines.append('export const STORES = %s' % json.dumps([
        {'key': k, 'name': n, 'district': d} for _, k, n, d in STORE_MAP
    ], ensure_ascii=False))
    lines.append('')
    lines.append('export const MONTHS = %s' % json.dumps([
        {'key': m, 'label': '%s年%02d月' % (m[:4], int(m[5:7]))} for m, _ in MONTH_FILES
    ], ensure_ascii=False))
    lines.append('')
    lines.append('// 每日明细（按 月份 -> 门店 -> 日期数组）')
    lines.append('export const DAILY = %s' % json.dumps(all_daily, ensure_ascii=False))
    lines.append('')
    lines.append('// 月度汇总（按 门店 -> 月份）')
    lines.append('export const MONTHLY = %s' % json.dumps(monthly, ensure_ascii=False))
    lines.append('')
    lines.append('// 菜品销售明细（按 月份 -> 门店 -> 菜品数组，按销售额降序）')
    lines.append('export const PRODUCTS = %s' % json.dumps(products, ensure_ascii=False))
    lines.append('')
    lines.append('// 员工绩效（薪资表 5 周合计）')
    lines.append('export const EMPLOYEES = %s' % json.dumps(emp_out, ensure_ascii=False))
    lines.append('')
    lines.append('// 员工月度薪资（按周拆分：6月2天 / 7月31天 / 8月2天）')
    lines.append('export const EMPLOYEE_MONTHLY = %s' % json.dumps(emp_monthly, ensure_ascii=False))
    lines.append('')
    lines.append('export const EMPLOYEE_MONTHS = %s' % json.dumps(sorted(emp_monthly.keys()), ensure_ascii=False))

    out = '\n'.join(lines) + '\n'
    out_path = os.path.abspath(OUT_PATH)
    with io.open(out_path, 'w', encoding='utf-8') as f:
        f.write(out)

    print('== 月度汇总校验（全部门店） ==')
    for mk, _ in MONTH_FILES:
        t = {'rev': 0.0, 'inc': 0.0, 'ord': 0, 'dish': 0}
        for key in SOURCE_TO_KEY.values():
            a = monthly[key][mk]
            t['rev'] += a['rev']
            t['inc'] += a['inc']
            t['ord'] += a['ord']
            t['dish'] += a['dish']
        print('%s: 营业额=%.2f 收入=%.2f 订单=%d 菜品销量=%d' % (mk, t['rev'], t['inc'], t['ord'], t['dish']))
    print('== 菜品明细校验（每店 TOP3） ==')
    for mk, fname in PRODUCT_FILES:
        for key in SOURCE_TO_KEY.values():
            plist = products[mk].get(key, [])
            top = ' | '.join('%s(销量%.0f/额%.0f)' % (p['name'][:10], p['sales'], p['amount']) for p in plist[:3])
            print('%s %s: %d 个菜品 | %s' % (mk, key, len(plist), top))
    print('== 员工（按工资排序） ==')
    for e in emp_out:
        print('  %s | %s | 工资=%.2f | 工时=%.1fh | 业绩提成=%.0f | 当班营业额=%.2f | 出勤=%d天' % (
            e['name'], e['storeName'], e['salary'], e['baseHours'] + e['otHours'], e['perf'], e['workedRevenue'], e['workedDays']))
    print('生成: %s' % out_path)


if __name__ == '__main__':
    main()